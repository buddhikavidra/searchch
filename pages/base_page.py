from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import openpyxl
import time
from openpyxl import Workbook

class BasePage():
    def __init__(self, browser, url, timeout=1):
        self.browser = browser
        self.url = url
        self.browser.implicitly_wait(timeout)

    def open(self):
        return self.browser.get(self.url)

    def is_element_present(self, how, what):
        time.sleep(3)
        try:
            self.browser.find_element(how, what)
        except NoSuchElementException:
            return False
        return True

    def is_not_element_present(self, how, what, timeout=4):
        try:
            WebDriverWait(self.browser, timeout).until(EC.presence_of_element_located((how, what)))
        except TimeoutException:
            return True

        return 
        
    def is_disappeared(self, how, what, timeout=4):
        try:
            WebDriverWait(self.browser, timeout, 1, TimeoutException).\
                until_not(EC.presence_of_element_located((how, what)))
        except TimeoutException:
            return False

        return True


    def array_to_csv_string(self,array):
        csv_string = ",".join(str(element) for element in array)
        return csv_string


    def write_array_to_excel(self, array, filename, sheet_name):
        try:
            workbook = openpyxl.load_workbook(filename)
        except FileNotFoundError:
            workbook = Workbook()
        
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            last_row = sheet.max_row + 1
        else:
            sheet = workbook.create_sheet(title=sheet_name)
            last_row = 1

        for col, value in enumerate(array, start=1):
            sheet.cell(row=last_row, column=col).value = value

        workbook.save(filename)
        print("Successfully appended")

    def current_url(self):
        try:
            current_url = self.browser.current_url
        except :
            current_url = ""
        return current_url